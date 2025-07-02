from __future__ import annotations

import openpyxl, pandas as pd
from pathlib import Path
from navigator import _select_show_all
from scraper   import _row_bg
from datetime import datetime

from config import (
    TARGET_FIELDS,         # dict like {"AdmissionDate": {...}, "Habitation": {...}, …}
    DATE_DEFAULTS,         # { 11: "09062025", 12: "09062024" }
    UPDATE_BTN,
    STUDENTS_XLSX,
    BASELINE_XLSX,
    TABLE_SEL, EDIT_ICON_SEL,
)

# ─────────────────────────────  baseline  ──────────────────────────────
def _load_baseline() -> dict[str, dict]:
    try:
        df = pd.read_excel(BASELINE_XLSX)
        df.columns = [c.lower().strip() for c in df.columns]
        return {str(r.uid): r._asdict() for r in df.itertuples(index=False)}
    except FileNotFoundError:
        print("⚠️  baseline XLSX missing – will only inspect, never update.")
        return {}

BASELINE = _load_baseline()

# ────────────────────────────── helpers ────────────────────────────────
def _force_type(elem, text: str):
    """Simulates full user input: clear + slow typing + blur."""
    elem.click()
    elem.press("Control+A")
    # elem.press("Backspace")
    elem.type(text, delay=80)  # slower = more reliable
    # elem.evaluate("e => e.blur()")  # trigger any on-blur validators    # slow enough for on-input scripts



def _stamp_comment(uid: str, std: int | str, note: str):
    wb_path = Path(STUDENTS_XLSX)
    wb = openpyxl.load_workbook(wb_path)
    ws = wb[f"Std{std}"]

    headers = [h.value for h in ws[1]]
    if "comment" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="comment")
        headers.append("comment")

    uid_col = headers.index("uid") + 1
    cmt_col = headers.index("comment") + 1

    for row in ws.iter_rows(min_row=2):
        if str(row[uid_col - 1].value).strip() == uid:
            row[cmt_col - 1].value = note
            break
    wb.save(wb_path)
def _find_first_white_row(page, std: int):
    wb_path = Path(STUDENTS_XLSX)
    wb = openpyxl.load_workbook(wb_path)
    ws = wb[f"Std{std}"]

    headers = [h.value for h in ws[1]]
    uid_col = headers.index("uid") + 1
    cmt_col = headers.index("comment") + 1 if "comment" in headers else None

    for tr in page.query_selector_all(f"{TABLE_SEL} > tbody > tr"):
        if _row_bg(tr) != "white":
            continue

        uid = tr.query_selector("td").inner_text().strip()

        for row in ws.iter_rows(min_row=2):
            if str(row[uid_col - 1].value).strip() == uid:
                comment = row[cmt_col - 1].value if cmt_col else ""
                if not comment:  # ✅ only return if comment is empty
                    return tr
                break

    return None

# ───────────────────────────── main worker ─────────────────────────────
def process_first_pending(page, std):
    """Edit every *white* row for this standard, until none remain."""
    if std == 9:
        print("⏭ Std-9 skipped by config.")
        return

    while True:
        # row = _find_first_white_row(page)
        row = _find_first_white_row(page, std)

        if row is None:
            print(f"✅ Std-{std}: nothing pending.")
            return

        uid = row.query_selector("td").inner_text().strip()
        print(f"✏  UID {uid}")

        # open edit-form
        row.query_selector(EDIT_ICON_SEL).click()
        page.wait_for_load_state("networkidle")

        baseline_row = BASELINE.get(uid, {})
        comments     = []

        for nice_name, spec in TARGET_FIELDS.items():
            html_id = spec["id"]
            selector = f"#{html_id}"

            # locate – allow for re-rendering
            try:
                elem = page.wait_for_selector(selector, state="attached", timeout=15_000)
            except Exception:
                print(f"❓ [{uid}] {selector} not found – skipping")
                comments.append(f"{nice_name}=NOT-FOUND")
                continue

            tag      = elem.evaluate("e=>e.tagName").lower()
            current  = elem.input_value().strip()
            desired  = str(baseline_row.get(nice_name.lower(), "")).strip()

            # class-specific default for Admission-date
            if not desired and html_id.endswith("TxtDateOfAddmission"):
                desired = DATE_DEFAULTS.get(std, "")

            if elem.is_disabled():
                print(f"🔒 [{uid}] {nice_name} is disabled – skipping")
                comments.append(f"{nice_name}=LOCKED")
                continue

            if html_id.endswith("TxtDateOfAddmission") and desired and desired != current:
                page.evaluate(
                    """(id)=> {
                        const el = document.getElementById(id);
                        el.value = '';
                        el.dispatchEvent(new Event('change', {bubbles:true}));
                    }""", html_id
                )
                _force_type(elem, desired)
                page.wait_for_timeout(1000)  # give JS time to validate

                # recheck after typing
                post_type = elem.input_value().strip()
                try:
                    # Normalize both to datetime for better matching
                    expected_dt = datetime.strptime(desired, "%d%m%Y")
                    actual_dt = datetime.strptime(post_type, "%d/%m/%Y")

                    if expected_dt != actual_dt:
                        comments.append(f"{nice_name}=FAILED({current}→{desired}↛{post_type})")
                        break
                    else:
                        comments.append(f"{nice_name}: {current}➜{post_type}")
                        continue
                except Exception as e:
                    comments.append(f"{nice_name}=DATE-PARSE-ERROR: {e}")
                    break

            # ---------- update if required ------------------------------------------
            if desired and desired != current:
                if tag == "select":
                    elem.select_option(desired)
                    page.wait_for_load_state("networkidle")
                else:
                    _force_type(elem, desired)
                comments.append(f"{nice_name}: {current}➜{desired}")
            else:
                if not current:
                    print(f"⚠️  [{uid}] {nice_name} is BLANK but baseline also empty – treated as OK")
                else:
                    print(f"👌 [{uid}] {nice_name} already correct")
                comments.append(f"{nice_name}=OK")

        # give any on-change scripts time to settle
        page.wait_for_timeout(5_000)

        # save
        try:
            print(f"💾 [{uid}] Clicking update button...")
            page.click(UPDATE_BTN)
            page.wait_for_load_state("networkidle")
        except Exception as e:
            comments.append(f"❌ Update button error: {e}")

        # log to Excel
        _stamp_comment(uid, std, "; ".join(comments))

        # table collapses after save – expand again
        _select_show_all(page)
